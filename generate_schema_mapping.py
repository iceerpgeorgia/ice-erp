"""
Generate Excel file with Prisma schema table and column mappings.
Structure: [[table1, table2, ...], [column1, column2, ...]]
"""

import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def parse_prisma_schema(schema_path):
    """Parse Prisma schema and extract models with their fields."""
    with open(schema_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    models = []
    
    # Find all model definitions
    model_pattern = r'model\s+(\w+)\s*\{([^}]+)\}'
    for match in re.finditer(model_pattern, content, re.MULTILINE | re.DOTALL):
        model_name = match.group(1)
        model_body = match.group(2)
        
        # Extract fields
        fields = []
        field_pattern = r'^\s+(\w+)\s+(\w+)'
        for field_match in re.finditer(field_pattern, model_body, re.MULTILINE):
            field_name = field_match.group(1)
            # Skip special directives
            if not field_name.startswith('@'):
                fields.append(field_name)
        
        # Extract @@map for table name
        map_match = re.search(r'@@map\("([^"]+)"\)', model_body)
        db_table_name = map_match.group(1) if map_match else model_name.lower()
        
        # Extract @map for field mappings
        field_mappings = {}
        for field in fields:
            field_map_pattern = rf'{field}\s+[^\n]+@map\("([^"]+)"\)'
            field_map_match = re.search(field_map_pattern, model_body)
            if field_map_match:
                field_mappings[field] = field_map_match.group(1)
            else:
                # Check if field is naturally snake_case
                field_mappings[field] = field
        
        models.append({
            'model_name': model_name,
            'db_table_name': db_table_name,
            'fields': fields,
            'field_mappings': field_mappings
        })
    
    return models

def create_mapping_excel(models, output_path):
    """Create Excel file with table and column mappings."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Schema Mapping"
    
    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    table_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Headers
    ws['A1'] = "Prisma Model"
    ws['B1'] = "Database Table"
    ws['C1'] = "TypeScript Field"
    ws['D1'] = "Database Column"
    
    for cell in ['A1', 'B1', 'C1', 'D1']:
        ws[cell].fill = header_fill
        ws[cell].font = header_font
        ws[cell].alignment = Alignment(horizontal='center', vertical='center')
    
    # Data
    row = 2
    for model in sorted(models, key=lambda x: x['db_table_name']):
        start_row = row
        
        for i, field in enumerate(model['fields']):
            db_column = model['field_mappings'][field]
            
            # Write model and table name only on first field
            if i == 0:
                ws[f'A{row}'] = model['model_name']
                ws[f'B{row}'] = model['db_table_name']
                ws[f'A{row}'].fill = table_fill
                ws[f'B{row}'].fill = table_fill
            
            ws[f'C{row}'] = field
            ws[f'D{row}'] = db_column
            
            row += 1
        
        # Merge model and table cells if multiple fields
        if len(model['fields']) > 1:
            ws.merge_cells(f'A{start_row}:A{row-1}')
            ws.merge_cells(f'B{start_row}:B{row-1}')
            ws[f'A{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'B{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    
    wb.save(output_path)
    print(f"\n✅ Created: {output_path}")
    print(f"   Models: {len(models)}")
    print(f"   Total fields: {sum(len(m['fields']) for m in models)}")

if __name__ == "__main__":
    schema_path = "prisma/schema.prisma"
    output_path = "schema_mapping.xlsx"
    
    print("Parsing Prisma schema...")
    models = parse_prisma_schema(schema_path)
    
    print(f"Found {len(models)} models")
    
    create_mapping_excel(models, output_path)
    
    # Print summary
    print("\n📊 Sample mappings:")
    for model in models[:3]:
        print(f"\n{model['model_name']} → {model['db_table_name']}")
        for field in list(model['fields'])[:5]:
            db_col = model['field_mappings'][field]
            if field != db_col:
                print(f"  {field} → {db_col}")
            else:
                print(f"  {field}")
