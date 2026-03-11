from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


def parse_prisma_tables(schema_text: str) -> list[dict[str, str]]:
    model_pattern = re.compile(r"model\s+(\w+)\s*\{(.*?)\n\}", re.DOTALL)
    map_pattern = re.compile(r"@@map\(\"([^\"]+)\"\)")

    rows: list[dict[str, str]] = []

    for model_name, model_body in model_pattern.findall(schema_text):
        table_name = model_name
        map_match = map_pattern.search(model_body)
        if map_match:
            table_name = map_match.group(1)

        has_insider = (
            "insider_uuid" in model_body
            or "insiderUuid" in model_body
            or "@map(\"insider_uuid\")" in model_body
        )

        rows.append(
            {
                "source": "prisma",
                "table_name": table_name,
                "model_name": model_name,
                "has_insider_uuid": "yes" if has_insider else "no",
                "decision": "",
                "notes": "",
            }
        )

    return sorted(rows, key=lambda r: r["table_name"].lower())


def build_workbook(rows: list[dict[str, str]], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "insider_policy"

    headers = [
        "source",
        "table_name",
        "model_name",
        "has_insider_uuid",
        "decision",
        "notes",
    ]
    ws.append(headers)

    for row in rows:
        ws.append([row[h] for h in headers])

    # Add known non-Prisma operational tables that are still part of business flow.
    extra_rows = [
        ["operational", "bank_transaction_batches", "", "no", "", "Batch partitions and BTC resolution"],
        ["operational", "conversion", "", "no", "", "Conversion parent rows"],
        ["operational", "conversion_entries", "", "no", "", "Conversion IN/OUT/FEE rows"],
        ["source_raw", "bog_gel_deconsolidated", "", "unknown", "", "Raw/deconsolidated bank source"],
        ["source_raw", "bog_usd_deconsolidated", "", "unknown", "", "Raw/deconsolidated bank source"],
        ["source_raw", "bog_eur_deconsolidated", "", "unknown", "", "Raw/deconsolidated bank source"],
        ["source_raw", "bog_aed_deconsolidated", "", "unknown", "", "Raw/deconsolidated bank source"],
        ["source_raw", "bog_gbp_deconsolidated", "", "unknown", "", "Raw/deconsolidated bank source"],
        ["source_raw", "bog_kzt_deconsolidated", "", "unknown", "", "Raw/deconsolidated bank source"],
        ["source_raw", "bog_cny_deconsolidated", "", "unknown", "", "Raw/deconsolidated bank source"],
        ["source_raw", "bog_try_deconsolidated", "", "unknown", "", "Raw/deconsolidated bank source"],
        ["source_raw", "tbc_gel_deconsolidated", "", "unknown", "", "Raw/deconsolidated bank source"],
    ]
    for extra in extra_rows:
        ws.append(extra)

    # Header style
    header_fill = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Decision dropdown (editable in Excel)
    last_row = ws.max_row
    decision_validation = DataValidation(
        type="list",
        formula1='"required,optional,no-insider"',
        allow_blank=True,
        showDropDown=True,
    )
    ws.add_data_validation(decision_validation)
    decision_validation.add(f"E2:E{last_row}")

    # Formatting
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{last_row}"

    widths = {
        "A": 14,
        "B": 42,
        "C": 28,
        "D": 17,
        "E": 16,
        "F": 55,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=1, max_col=6):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(output_path)


def main() -> None:
    repo_root = Path(__file__).resolve().parents[1]
    schema_path = repo_root / "prisma" / "schema.prisma"
    output_path = repo_root / "docs" / "insider_uuid_table_policy.xlsx"

    schema_text = schema_path.read_text(encoding="utf-8")
    rows = parse_prisma_tables(schema_text)
    build_workbook(rows, output_path)

    print(f"Created: {output_path}")
    print(f"Rows exported: {len(rows)} Prisma tables + operational/source rows")


if __name__ == "__main__":
    main()
