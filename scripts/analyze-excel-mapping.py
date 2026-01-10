import openpyxl
import json

def analyze_excel_mapping(file_path):
    """
    Analyze the Excel file with column mapping and data.
    """
    wb = openpyxl.load_workbook(file_path)
    
    print(f"📄 Analyzing: {file_path}")
    print(f"📋 Sheets: {wb.sheetnames}\n")
    
    # Analyze each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        print(f"\n{'='*80}")
        print(f"📊 Sheet: {sheet_name}")
        print(f"{'='*80}")
        
        # Get dimensions
        print(f"Rows: {ws.max_row}, Columns: {ws.max_column}")
        
        # Get headers (first row)
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            headers.append(cell_value)
        
        print(f"\n📌 Headers ({len(headers)} columns):")
        for idx, header in enumerate(headers, 1):
            print(f"   {idx}. {header}")
        
        # Show first 3 data rows
        if ws.max_row > 1:
            print(f"\n📝 Sample Data (first 3 rows):")
            for row_num in range(2, min(5, ws.max_row + 1)):
                print(f"\n   Row {row_num}:")
                row_data = {}
                for col_idx, header in enumerate(headers, 1):
                    cell_value = ws.cell(row=row_num, column=col_idx).value
                    if cell_value is not None:
                        row_data[header if header else f"Col{col_idx}"] = cell_value
                
                for key, value in row_data.items():
                    value_str = str(value)[:60]  # Truncate long values
                    print(f"      {key}: {value_str}")
    
    wb.close()

if __name__ == "__main__":
    file_path = "templates/GE78BG0000000893486000GEL.xlsx"
    analyze_excel_mapping(file_path)
