#!/usr/bin/env python3
"""
Re-import counteragent account numbers from Excel file
Matches directly against consolidated_bank_accounts table
"""

import openpyxl
import psycopg2
import os

# Database connection
def get_database_url():
    try:
        with open('.env.local', 'r', encoding='utf-8') as f:
            for line in f:
                if line.startswith('DATABASE_URL='):
                    url = line.split('=', 1)[1].strip().strip('"').strip("'")
                    if '?schema=' in url:
                        url = url.split('?schema=')[0]
                    return url
    except Exception as e:
        print(f"❌ Error reading .env.local: {e}")
    return None

def convert_from_scientific(value):
    """Convert scientific notation to regular number string"""
    if not value:
        return None
    value_str = str(value).strip()
    
    # Skip if it starts with 'GE' (Georgian bank account format)
    if value_str.startswith('GE'):
        return value_str
    
    # Check for scientific notation (but only if it's actually numeric)
    if 'e' in value_str.lower() and not value_str.startswith('GE'):
        try:
            numeric_value = float(value_str)
            int_value = int(numeric_value)
            return str(int_value)
        except (ValueError, OverflowError):
            return value_str
    return value_str

def main():
    print("\n" + "="*80)
    print("RE-IMPORT COUNTERAGENT ACCOUNTS FROM EXCEL")
    print("="*80 + "\n")

    # Connect to database
    DATABASE_URL = get_database_url()
    if not DATABASE_URL:
        print("❌ Could not get DATABASE_URL")
        return

    try:
        conn = psycopg2.connect(DATABASE_URL)
        cursor = conn.cursor()
        print("✅ Connected to database\n")
    except Exception as e:
        print(f"❌ Database connection failed: {e}")
        return

    # Load Excel file
    excel_path = 'templates/GE78BG0000000893486000GEL.xlsx'
    print(f"📂 Loading Excel file: {excel_path}")
    
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        print(f"✅ Loaded Excel file\n")
    except Exception as e:
        print(f"❌ Could not load Excel file: {e}")
        conn.close()
        return

    # Excel file is transposed - headers are in column B
    # We need to find which row has which data
    ref_row = None
    entries_id_row = None
    sender_account_row = None
    benef_account_row = None
    debit_row = None

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=100, max_col=2, values_only=True), start=1):
        header = row[1] if len(row) > 1 else None
        if header == 'Ref':
            ref_row = row_idx
        elif header == 'ოპერაციის იდ':
            entries_id_row = row_idx
        elif header == 'გამგზავნის ანგარიშის ნომერი':
            sender_account_row = row_idx
        elif header == 'ბენეფიციარის ანგარიშის ნომერი':
            benef_account_row = row_idx
        elif header == 'დებეტი':
            debit_row = row_idx
    
    print(f"📋 Found row mappings:")
    print(f"   Ref: row {ref_row}")
    print(f"   ოპერაციის იდ: row {entries_id_row}")
    print(f"   გამგზავნის ანგარიშის ნომერი: row {sender_account_row}")
    print(f"   ბენეფიციარის ანგარიშის ნომერი: row {benef_account_row}")
    print(f"   დებეტი: row {debit_row}")
    print()

    if None in [ref_row, entries_id_row, sender_account_row, benef_account_row, debit_row]:
        print(f"❌ Could not find required rows")
        conn.close()
        return

    print("✅ Found all required rows\n")
    print("="*80)
    print("PROCESSING RECORDS")
    print("="*80 + "\n")

    # Get max column to know how many records there are
    max_col = ws.max_column
    print(f"   Total records to process: {max_col - 2}\n")

    # Process columns (each column starting from C is a record)
    updated_count = 0
    not_found_count = 0
    no_account_count = 0
    error_count = 0

    for col_idx in range(3, max_col + 1):  # Start from column C (index 3)
        if (col_idx - 2) % 1000 == 0:
            print(f"   Processing record {col_idx - 2}...")

        try:
            doc_key = ws.cell(row=ref_row, column=col_idx).value
            entries_id = ws.cell(row=entries_id_row, column=col_idx).value
            sender_account = ws.cell(row=sender_account_row, column=col_idx).value
            benef_account = ws.cell(row=benef_account_row, column=col_idx).value
            debit = ws.cell(row=debit_row, column=col_idx).value

            if not doc_key or not entries_id:
                continue

            # Convert scientific notation if present
            sender_account = convert_from_scientific(sender_account)
            benef_account = convert_from_scientific(benef_account)

            # Determine which account to use based on direction
            if debit is None or debit == '' or debit == 0:
                # Incoming payment - use sender's account
                counteragent_account = sender_account
            else:
                # Outgoing payment - use beneficiary's account
                counteragent_account = benef_account

            if not counteragent_account:
                no_account_count += 1
                continue

            # Update consolidated_bank_accounts directly
            cursor.execute("""
                UPDATE consolidated_bank_accounts
                SET counteragent_account_number = %s
                WHERE doc_key = %s
                AND entries_id = %s::text
                AND (counteragent_account_number IS NULL 
                     OR counteragent_account_number LIKE '%%e+%%'
                     OR counteragent_account_number LIKE '%%e-%%')
            """, (counteragent_account, doc_key, str(entries_id)))

            if cursor.rowcount > 0:
                updated_count += 1
                if updated_count <= 10:
                    print(f"   ✅ Updated DocKey={doc_key}, EntriesId={entries_id} → {counteragent_account}")
            else:
                not_found_count += 1

        except Exception as e:
            error_count += 1
            if error_count <= 5:
                print(f"   ❌ Error processing column {col_idx}: {e}")

    # Commit changes
    conn.commit()
    cursor.close()
    conn.close()

    print("\n" + "="*80)
    print("SUMMARY")
    print("="*80)
    print(f"  ✅ Updated: {updated_count}")
    print(f"  ⚠️  Not found or already has value: {not_found_count}")
    print(f"  ℹ️  No account in Excel: {no_account_count}")
    print(f"  ❌ Errors: {error_count}")
    print("="*80 + "\n")

    print("✅ Re-import completed successfully!")

if __name__ == '__main__':
    main()
