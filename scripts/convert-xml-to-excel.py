import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os

def xml_to_excel(xml_file, output_file):
    """
    Convert BOG XML statement to Excel format for reference.
    """
    
    # Parse XML
    tree = ET.parse(xml_file)
    root = tree.getroot()
    
    # Define namespace
    ns = {'gemini': 'http://www.mygemini.com/schemas/mygemini'}
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOG_Statement_Reference"
    
    # Define columns (matching the template structure)
    columns = [
        'dockey', 'entriesid', 'cancopydocument', 'canviewdocument', 'canprintdocument',
        'isreval', 'docnomination', 'docinformation', 'docsrcamt', 'docsrcccy',
        'docdstamt', 'docdstccy', 'docrecdate', 'docbranch', 'docdepartment',
        'docprodgroup', 'docno', 'docvaluedate', 'docsendername', 'docsenderinn',
        'docsenderacctno', 'docsenderbic', 'docsenderbicname', 'docbenefname',
        'docbenefinn', 'docbenefacctno', 'docbenefbic', 'docbenefbicname',
        'docpayername', 'docpayerinn', 'docactualdate', 'doccoracct', 'doccorbic',
        'doccorbankname', 'doccomment', 'ccyrate', 'entrypdate', 'entrydocno',
        'entrylacct', 'entrylacctold', 'entrydbamt', 'entrydbamtbase', 'entrycramt',
        'entrycramtbase', 'outbalance', 'entryamtbase', 'entrycomment',
        'entrydepartment', 'entryacctpoint'
    ]
    
    # Set column headers with styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.value = col_name.upper()
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(idx)].width = 15
    
    # Extract data from XML
    row_num = 2
    
    # Helper function to get text from XML element
    def get_text(element, tag):
        child = element.find(tag)
        return child.text if child is not None and child.text else ''
    
    # Process DETAIL elements directly (no Document or Entries wrapper)
    details = root.find('.//DETAILS')
    if details is None:
        print("❌ No DETAILS section found in XML")
        wb.save(output_file)
        return 0
    
    for detail in details.findall('DETAIL'):
        DocKey = get_text(detail, 'DocKey')
        EntriesId = get_text(detail, 'EntriesId')
        
        # Skip records without DocKey and EntriesId
        if not DocKey or not EntriesId:
            continue
        
        # Build row data - all fields are directly under DETAIL
        row_data = {
            'dockey': DocKey,
            'entriesid': EntriesId,
            'cancopydocument': get_text(detail, 'CanCopyDocument'),
            'canviewdocument': get_text(detail, 'CanViewDocument'),
            'canprintdocument': get_text(detail, 'CanPrintDocument'),
            'isreval': get_text(detail, 'IsReval'),
            'docnomination': get_text(detail, 'DocNomination'),
            'docinformation': get_text(detail, 'DocInformation'),
            'docsrcamt': get_text(detail, 'DocSrcAmt'),
            'docsrcccy': get_text(detail, 'DocSrcCcy'),
            'docdstamt': get_text(detail, 'DocDstAmt'),
            'docdstccy': get_text(detail, 'DocDstCcy'),
            'docrecdate': get_text(detail, 'DocRecDate'),
            'docbranch': get_text(detail, 'DocBranch'),
            'docdepartment': get_text(detail, 'DocDepartment'),
            'docprodgroup': get_text(detail, 'DocProdGroup'),
            'docno': get_text(detail, 'DocNo'),
            'docvaluedate': get_text(detail, 'DocValueDate'),
            'docsendername': get_text(detail, 'DocSenderName'),
            'docsenderinn': get_text(detail, 'DocSenderInn'),
            'docsenderacctno': get_text(detail, 'DocSenderAcctNo'),
            'docsenderbic': get_text(detail, 'DocSenderBic'),
            'docsenderbicname': get_text(detail, 'DocSenderBicName'),
            'docbenefname': get_text(detail, 'DocBenefName'),
            'docbenefinn': get_text(detail, 'DocBenefInn'),
            'docbenefacctno': get_text(detail, 'DocBenefAcctNo'),
            'docbenefbic': get_text(detail, 'DocBenefBic'),
            'docbenefbicname': get_text(detail, 'DocBenefBicName'),
            'docpayername': get_text(detail, 'DocPayerName'),
            'docpayerinn': get_text(detail, 'DocPayerInn'),
            'docactualdate': get_text(detail, 'DocActualDate'),
            'doccoracct': get_text(detail, 'DocCorAcct'),
            'doccorbic': get_text(detail, 'DocCorBic'),
            'doccorbankname': get_text(detail, 'DocCorBankName'),
            'doccomment': get_text(detail, 'DocComment'),
            'ccyrate': get_text(detail, 'CcyRate'),
            'entrypdate': get_text(detail, 'EntryPDate'),
            'entrydocno': get_text(detail, 'EntryDocNo'),
            'entrylacct': get_text(detail, 'EntryLAcct'),
            'entrylacctold': get_text(detail, 'EntryLAcctOld'),
            'entrydbamt': get_text(detail, 'EntryDbAmt'),
            'entrydbamtbase': get_text(detail, 'EntryDbAmtBase'),
            'entrycramt': get_text(detail, 'EntryCrAmt'),
            'entrycramtbase': get_text(detail, 'EntryCrAmtBase'),
            'outbalance': get_text(detail, 'OutBalance'),
            'entryamtbase': get_text(detail, 'EntryAmtBase'),
            'entrycomment': get_text(detail, 'EntryComment'),
            'entrydepartment': get_text(detail, 'EntryDepartment'),
            'entryacctpoint': get_text(detail, 'EntryAcctPoint'),
        }
        
        # Write row to Excel
        for idx, col_name in enumerate(columns, start=1):
            ws.cell(row=row_num, column=idx).value = row_data.get(col_name, '')
        
        row_num += 1
    
    # Freeze top row
    ws.freeze_panes = "A2"
    
    # Auto-filter
    ws.auto_filter.ref = ws.dimensions
    
    # Save workbook
    wb.save(output_file)
    
    records_count = row_num - 2
    print(f"✅ Converted {records_count} records from XML to Excel")
    print(f"📄 Output file: {output_file}")
    
    return records_count

if __name__ == "__main__":
    xml_file = "Statement_206598021.xml"
    output_file = "templates/Statement_206598021_reference.xlsx"
    
    if not os.path.exists(xml_file):
        print(f"❌ XML file not found: {xml_file}")
        exit(1)
    
    xml_to_excel(xml_file, output_file)
