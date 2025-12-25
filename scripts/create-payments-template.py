#!/usr/bin/env python3
"""
Create a user-friendly payments import template
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Create template with example data
data = {
    'payment_id': ['Example: a3f5c9_4b_12d8 (optional - auto-generated if empty)', ''],
    'project_uuid': ['Example: 123e4567-e89b-12d3-a456-426614174000', ''],
    'counteragent_uuid': ['Example: 123e4567-e89b-12d3-a456-426614174001', ''],
    'financial_code_uuid': ['Example: 123e4567-e89b-12d3-a456-426614174002', ''],
    'job_uuid': ['Example: 123e4567-e89b-12d3-a456-426614174003', ''],
    'income_tax': ['true or false', ''],
    'currency_uuid': ['Example: 123e4567-e89b-12d3-a456-426614174004', ''],
    'is_active': ['true (default: true if empty)', '']
}

df = pd.DataFrame(data)

# Write to Excel
excel_path = 'templates/payments_import_template.xlsx'
with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='Payments')

# Format the Excel file
workbook = load_workbook(excel_path)
worksheet = workbook['Payments']

# Style header row
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF')

for cell in worksheet[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Auto-adjust column widths
for column in worksheet.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)
    worksheet.column_dimensions[column_letter].width = adjusted_width

# Style example row
example_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
for cell in worksheet[2]:
    cell.fill = example_fill
    cell.font = Font(italic=True, color='666666')

workbook.save(excel_path)

print(f'✓ Created enhanced payments import template: {excel_path}')
print('\nColumns:')
for col in df.columns:
    print(f'  • {col}')
print('\nNotes:')
print('  • payment_id: optional - provide to import from Google Sheets, auto-generated if empty')
print('  • payment_id format: 6hex_2hex_4hex (e.g., a3f5c9_4b_12d8)')
print('  • record_uuid: always auto-generated')
print('  • income_tax: boolean (true/false)')
print('  • is_active: defaults to true if not provided')
print('  • All UUIDs must reference existing records in respective tables')
