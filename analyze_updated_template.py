import openpyxl
from openpyxl.utils import get_column_letter
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('handover template.xlsx', data_only=False)
ws = wb['Handover']

print("="*160)
print("UPDATED TEMPLATE - ALL CELLS WITH CONTENT")
print("="*160 + "\n")

for row_num in range(1, min(10, ws.max_row + 1)):
    print(f"\n--- ROW {row_num} ---")
    for col_num in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell_ref = f"{get_column_letter(col_num)}{row_num}"
        
        if cell.value is not None:
            value = cell.value
            
            if hasattr(cell, 'data_type') and cell.data_type == 'f':
                if hasattr(value, 'text'):
                    formula_text = value.text
                else:
                    formula_text = str(value)
                print(f"  {cell_ref:<5} [FORMULA] {formula_text}")
            else:
                val_str = str(value)[:100]
                print(f"  {cell_ref:<5} [VALUE]   {val_str}")

print("\n\n" + "="*160)
print("MERGED CELLS (updated)")
print("="*160)
for merged_range in ws.merged_cells.ranges:
    print(f"  {merged_range}")

print("\n" + "="*160)
print("DATE CELL LOCATION")
print("="*160)
print("\nLooking for date references in formulas...")
for row_num in range(1, 20):
    for col_num in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_num, column=col_num)
        if hasattr(cell, 'data_type') and cell.data_type == 'f':
            if hasattr(cell.value, 'text'):
                formula_text = cell.value.text
            else:
                formula_text = str(cell.value)
            
            # Look for cell references like $V$3, etc.
            if '$' in formula_text and ':' not in formula_text:
                cell_ref = f"{get_column_letter(col_num)}{row_num}"
                print(f"  {cell_ref}: {formula_text}")
