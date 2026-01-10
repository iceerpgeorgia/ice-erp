import openpyxl

def extract_column_mapping(file_path):
    """
    Extract the full column mapping from the Suggested columns sheet.
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb['Suggested columns']
    
    print("📋 Column Mapping Analysis")
    print("="*80)
    print(f"{'Database Column':<30} | {'Your Excel Column':<30}")
    print("="*80)
    
    mappings = []
    for row_num in range(2, ws.max_row + 1):
        db_col = ws.cell(row=row_num, column=1).value
        excel_col = ws.cell(row=row_num, column=2).value
        
        if db_col:
            print(f"{str(db_col):<30} | {str(excel_col) if excel_col else '(not mapped)':<30}")
            mappings.append({
                'db_column': db_col,
                'excel_column': excel_col,
                'is_mapped': excel_col is not None and excel_col != ''
            })
    
    print("\n" + "="*80)
    print(f"\n📊 Summary:")
    mapped = sum(1 for m in mappings if m['is_mapped'])
    unmapped = len(mappings) - mapped
    print(f"   ✅ Mapped columns: {mapped}")
    print(f"   ⚠️  Unmapped columns: {unmapped}")
    
    # Analyze data columns
    print(f"\n📄 Your Historical Data Columns:")
    data_ws = wb['GE78BG0000000893486000GEL']
    data_headers = []
    for col in range(1, data_ws.max_column + 1):
        header = data_ws.cell(row=1, column=col).value
        if header:
            data_headers.append(header)
            # Check if this column is mapped
            mapped_to = None
            for m in mappings:
                if m['excel_column'] == header:
                    mapped_to = m['db_column']
                    break
            
            status = f"→ {mapped_to}" if mapped_to else "(not used)"
            print(f"   {col:2d}. {header:<40} {status}")
    
    wb.close()
    
    return mappings

if __name__ == "__main__":
    extract_column_mapping("templates/GE78BG0000000893486000GEL.xlsx")
