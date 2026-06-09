#!/usr/bin/env python3
import openpyxl
from openpyxl.utils import get_column_letter

# Load the workbook
wb = openpyxl.load_workbook('handover template.xlsx', data_only=False)

# List all sheet names
print("Sheet names:", wb.sheetnames)
print("\n" + "="*100)

# Analyze the 'Handover' sheet
if 'Handover' in wb.sheetnames:
    ws = wb['Handover']
    
    print(f"\nHANDOVER SHEET ANALYSIS")
    print(f"Dimensions: {ws.dimensions}")
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
    print(f"\n" + "="*100)
    
    print("\n📋 STRUCTURE - All cells with content:\n")
    
    for row_num in range(1, ws.max_row + 1):
        for col_num in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_num, column=col_num)
            if cell.value is not None:
                cell_ref = f"{get_column_letter(col_num)}{row_num}"
                value_str = str(cell.value)
                
                # Identify the type
                cell_type = "VALUE"
                if hasattr(cell, 'data_type') and cell.data_type == 'f':
                    cell_type = "FORMULA"
                elif value_str.startswith('='):
                    cell_type = "FORMULA"
                elif '{{' in value_str or '}}' in value_str:
                    cell_type = "PLACEHOLDER"
                elif '$' in value_str and ('{' in value_str or '}' in value_str):
                    cell_type = "PLACEHOLDER"
                
                # Get cell width for better formatting
                width = ws.column_dimensions[get_column_letter(col_num)].width or 12
                
                print(f"  [{cell_type:12}] {cell_ref:5} | {value_str[:80]}")

print("\n" + "="*100)
print("\n🔍 FORMULAS & PLACEHOLDERS ONLY:\n")

formulas = []
placeholders = []

for row_num in range(1, ws.max_row + 1):
    for col_num in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_num, column=col_num)
        if cell.value is not None:
            cell_ref = f"{get_column_letter(col_num)}{row_num}"
            value_str = str(cell.value)
            
            if hasattr(cell, 'data_type') and cell.data_type == 'f':
                formulas.append((cell_ref, value_str))
            elif value_str.startswith('='):
                formulas.append((cell_ref, value_str))
            elif '{{' in value_str or '}}' in value_str or ('$' in value_str and ('{' in value_str or '}' in value_str)):
                placeholders.append((cell_ref, value_str))

print("📐 FORMULAS:")
for ref, formula in formulas:
    print(f"  {ref}: {formula}")

print("\n📍 PLACEHOLDERS:")
for ref, placeholder in placeholders:
    print(f"  {ref}: {placeholder}")

print("\n" + "="*100)
print("\n🎯 SUMMARY:\n")
print(f"  Total cells: {sum(1 for row in ws.iter_rows() for cell in row if cell.value is not None)}")
print(f"  Formulas: {len(formulas)}")
print(f"  Placeholders: {len(placeholders)}")
print(f"  Content cells: {sum(1 for row in ws.iter_rows() for cell in row if cell.value is not None)}")

# Print merged cells if any
if ws.merged_cells:
    print(f"\n📦 MERGED CELLS: {len(ws.merged_cells)}")
    for merged_range in ws.merged_cells.ranges:
        print(f"  {merged_range}")

EOF
