"""
apply_wb_items_inventory_bindings.py

Reads inventory bindings from "RS.GE - WB_Items_IN.xlsx" and applies them to
the rs_waybills_in_items table in the database.

Matching key: (waybill_no, goods_name)  — goods_code (col 18) is skipped because
Excel misreads string codes like "08-20" as dates.

Fields written to rs_waybills_in_items:
  - inventory_uuid        (from col 42, "საქონელი GUID")
  - project_uuid          (from col 41, "პროექტი GUID")   — only when not null
  - financial_code_uuid   (from col 43, "კოდის GUID")     — only when not null

Usage:
  python apply_wb_items_inventory_bindings.py            # dry-run (no DB writes)
  python apply_wb_items_inventory_bindings.py --apply    # write to DB
"""

import sys
import urllib.parse
import re
from collections import defaultdict

import openpyxl
import psycopg2
import psycopg2.extras

EXCEL_PATH = "RS.GE - WB_Items_IN.xlsx"
DB_DIRECT_URL = (
    "postgresql://postgres.fojbzghphznbslqwurrm:fulebimojviT1985%25"
    "@aws-1-eu-west-1.pooler.supabase.com:5432/postgres"
)

DRY_RUN = "--apply" not in sys.argv

UUID_RE = re.compile(r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$', re.IGNORECASE)


def clean_uuid(val) -> str | None:
    """Return lowercase UUID string if valid, else None."""
    if not val:
        return None
    s = str(val).strip().lower()
    return s if UUID_RE.match(s) else None


# ---------------------------------------------------------------------------
# Step 1: Parse Excel bindings
# ---------------------------------------------------------------------------
print("=== Step 1: Reading Excel bindings ===")
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb["WB_Items_IN"]

# (waybill_no, goods_name) -> { inventory_uuid, project_uuid, financial_code_uuid }
binding_map: dict[tuple, dict] = {}
conflicts: list[tuple] = []        # keys with 2+ different inventory UUIDs
skipped_no_waybill = 0
skipped_no_name = 0

# Track duplicates for conflict detection
key_to_inventory: dict[tuple, set] = defaultdict(set)

all_bindings = []

for row_idx in range(2, ws.max_row + 1):
    row = [ws.cell(row_idx, c).value for c in range(1, 45)]

    inventory_uuid = row[41]   # col 42
    if not inventory_uuid:
        continue

    waybill_no_raw = row[24]   # col 25
    goods_name = row[18]       # col 19 — RS.GE goods name
    project_uuid = row[40]     # col 41
    financial_code_uuid = row[42]  # col 43

    if not waybill_no_raw:
        skipped_no_waybill += 1
        continue
    if not goods_name:
        skipped_no_name += 1
        continue

    waybill_no = str(waybill_no_raw).strip()
    goods_name = str(goods_name).strip()
    inventory_uuid = clean_uuid(inventory_uuid)
    if not inventory_uuid:
        continue

    key = (waybill_no, goods_name)
    key_to_inventory[key].add(inventory_uuid)
    all_bindings.append((key, inventory_uuid, project_uuid, financial_code_uuid))

# Identify conflicts (same key, different inventory UUIDs)
conflict_keys = {k for k, v in key_to_inventory.items() if len(v) > 1}
print(f"  Total bound rows in Excel:      {len(all_bindings)}")
print(f"  Skipped (no waybill_no):        {skipped_no_waybill}")
print(f"  Skipped (no goods_name):        {skipped_no_name}")
print(f"  Unique (waybill_no, goods_name) keys: {len(key_to_inventory)}")
print(f"  Keys with conflicting UUIDs:    {len(conflict_keys)}")
if conflict_keys:
    for k in list(conflict_keys)[:5]:
        print(f"    CONFLICT: {k} -> {key_to_inventory[k]}")

# Build final binding_map (skip conflicts, use first occurrence for duplicates)
for key, inventory_uuid, project_uuid, financial_code_uuid in all_bindings:
    if key in conflict_keys:
        continue  # skip ambiguous keys
    if key in binding_map:
        continue  # already added (same guid — safe to skip duplicate)
    binding_map[key] = {
        "inventory_uuid": inventory_uuid,
        "project_uuid": clean_uuid(project_uuid),
        "financial_code_uuid": clean_uuid(financial_code_uuid),
    }

print(f"  Final unique binding keys:      {len(binding_map)}")

# ---------------------------------------------------------------------------
# Step 2: Connect to DB
# ---------------------------------------------------------------------------
print("\n=== Step 2: Connecting to database ===")
parsed = urllib.parse.urlparse(DB_DIRECT_URL)
conn = psycopg2.connect(
    host=parsed.hostname,
    port=parsed.port,
    dbname=parsed.path[1:],
    user=parsed.username,
    password=urllib.parse.unquote(parsed.password),
    sslmode="require",
    connect_timeout=15,
)
conn.autocommit = False
cur = conn.cursor()
print("  Connected.")

# ---------------------------------------------------------------------------
# Step 3: Load existing items from DB (waybill_no + goods_name)
# ---------------------------------------------------------------------------
print("\n=== Step 3: Loading existing DB items ===")
cur.execute(
    "SELECT uuid, waybill_no, goods_name, inventory_uuid "
    "FROM rs_waybills_in_items "
    "WHERE waybill_no IS NOT NULL AND goods_name IS NOT NULL"
)
db_rows = cur.fetchall()
print(f"  DB rows with waybill_no+goods_name: {len(db_rows)}")

# Build DB lookup: (waybill_no, goods_name) -> list of (uuid, current_inventory_uuid)
db_lookup: dict[tuple, list] = defaultdict(list)
for uuid, waybill_no, goods_name, inv_uuid in db_rows:
    db_lookup[(waybill_no.strip(), goods_name.strip())].append((uuid, inv_uuid))

# ---------------------------------------------------------------------------
# Step 4a: Validate inventory_uuid values exist in inventories table
# ---------------------------------------------------------------------------
print("\n=== Step 4a: Validating inventory UUIDs ===")
all_inv_uuids = set(v["inventory_uuid"] for v in binding_map.values())
if all_inv_uuids:
    cur.execute(
        "SELECT uuid::text FROM inventories WHERE uuid::text = ANY(%s)",
        (list(all_inv_uuids),),
    )
    valid_uuids = set(str(row[0]).lower() for row in cur.fetchall())
else:
    valid_uuids = set()

invalid_uuids = all_inv_uuids - valid_uuids
print(f"  Total unique inventory UUIDs in Excel: {len(all_inv_uuids)}")
print(f"  Valid (exist in inventories table):    {len(valid_uuids)}")
print(f"  Invalid (not in inventories):          {len(invalid_uuids)}")
if invalid_uuids:
    print(f"  Sample invalid UUIDs:")
    for u in list(invalid_uuids)[:5]:
        print(f"    {u}")

# Filter binding_map to only valid inventory UUIDs
invalid_binding_keys = [k for k, v in binding_map.items() if v["inventory_uuid"] in invalid_uuids]
for k in invalid_binding_keys:
    del binding_map[k]
print(f"  Binding keys removed (invalid UUID):   {len(invalid_binding_keys)}")
print(f"  Remaining binding keys:                {len(binding_map)}")

# ---------------------------------------------------------------------------
# Step 4: Match and prepare updates
# ---------------------------------------------------------------------------
print("\n=== Step 4: Computing updates ===")

matched = 0
unmatched_keys = []
already_set = 0
update_rows = []  # (inventory_uuid, project_uuid, financial_code_uuid, uuid)

for key, binding in binding_map.items():
    db_items = db_lookup.get(key)
    if not db_items:
        unmatched_keys.append(key)
        continue

    for item_uuid, current_inv_uuid in db_items:
        if current_inv_uuid is not None:
            already_set += 1
            continue  # skip if already bound
        update_rows.append((
            binding["inventory_uuid"],
            binding["project_uuid"],
            binding["financial_code_uuid"],
            item_uuid,
        ))
        matched += 1

print(f"  Matched and to be updated:  {matched}")
print(f"  Already had inventory_uuid: {already_set}")
print(f"  Excel keys not in DB:       {len(unmatched_keys)}")
if unmatched_keys:
    print(f"  Sample unmatched keys:")
    for k in unmatched_keys[:5]:
        print(f"    {k}")

# ---------------------------------------------------------------------------
# Step 5: Apply updates (or print dry-run summary)
# ---------------------------------------------------------------------------
if DRY_RUN:
    print(f"\n=== DRY RUN: would update {matched} rows ===")
    print("  Re-run with --apply to write to DB.")
    print("\n  Sample updates (first 5):")
    for row in update_rows[:5]:
        print(f"    uuid={row[3]}  inv={row[0]}  proj={row[1]}  fc={row[2]}")
    conn.close()
    sys.exit(0)

print(f"\n=== Step 5: Applying {matched} updates ===")

BATCH_SIZE = 500
updated_total = 0

for i in range(0, len(update_rows), BATCH_SIZE):
    batch = update_rows[i : i + BATCH_SIZE]
    psycopg2.extras.execute_batch(
        cur,
        """
        UPDATE rs_waybills_in_items SET
            inventory_uuid = %s::uuid,
            project_uuid   = COALESCE(%s::uuid, project_uuid),
            financial_code_uuid = COALESCE(%s::uuid, financial_code_uuid),
            updated_at     = NOW()
        WHERE uuid = %s::uuid
          AND inventory_uuid IS NULL
        """,
        batch,
        page_size=BATCH_SIZE,
    )
    updated_total += len(batch)
    pct = int(updated_total / len(update_rows) * 100)
    print(f"  ... {updated_total}/{len(update_rows)} ({pct}%)")

conn.commit()
conn.close()

print(f"\n=== Done ===")
print(f"  Updated rows:           {updated_total}")
print(f"  Conflict keys skipped:  {len(conflict_keys)}")
print(f"  Excel keys not in DB:   {len(unmatched_keys)}")
print(f"  Already had binding:    {already_set}")
