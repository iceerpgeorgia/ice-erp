import openpyxl
from datetime import datetime

def compare_formats(historical_file, xml_reference_file):
    """
    Compare historical Excel format with XML-derived format to understand column mappings.
    """
    
    # Load both files
    wb_hist = openpyxl.load_workbook(historical_file)
    wb_xml = openpyxl.load_workbook(xml_reference_file)
    
    ws_hist = wb_hist['GE78BG0000000893486000GEL']
    ws_xml = wb_xml['BOG_Statement_Reference']
    
    print("="*100)
    print("COMPARING HISTORICAL DATA (2018) vs XML FORMAT (2025)")
    print("="*100)
    
    # Get headers
    hist_headers = [ws_hist.cell(1, col).value for col in range(1, ws_hist.max_column + 1)]
    xml_headers = [ws_xml.cell(1, col).value for col in range(1, ws_xml.max_column + 1)]
    
    print(f"\n📊 Historical Excel: {len(hist_headers)} columns")
    print(f"📊 XML Reference: {len(xml_headers)} columns")
    
    # Compare sample records
    print("\n" + "="*100)
    print("SAMPLE RECORD COMPARISON (First data row)")
    print("="*100)
    
    print("\n🔹 HISTORICAL FORMAT (Row 2):")
    print("-" * 100)
    for col_idx, header in enumerate(hist_headers, 1):
        if header:
            value = ws_hist.cell(2, col_idx).value
            print(f"   {header:<40} = {value}")
    
    print("\n🔹 XML FORMAT (Row 2):")
    print("-" * 100)
    for col_idx, header in enumerate(xml_headers, 1):
        if header:
            value = ws_xml.cell(2, col_idx).value
            if value:  # Only show non-empty
                print(f"   {header:<40} = {value}")
    
    # Find matching transaction (by amount or description)
    print("\n" + "="*100)
    print("MATCHING TRANSACTIONS ANALYSIS")
    print("="*100)
    
    # Get a few historical transactions
    hist_transaction = {
        'თარიღი': ws_hist.cell(2, 1).value,
        'Ref': ws_hist.cell(2, 9).value,
        'ოპერაციის იდ': ws_hist.cell(2, 8).value,
        'დებეტი': ws_hist.cell(2, 4).value,
        'კრედიტი': ws_hist.cell(2, 5).value,
        'ოპერაციის შინაარსი': ws_hist.cell(2, 6).value,
        'დანიშნულება': ws_hist.cell(2, 20).value,
        'დამატებითი ინფორმაცია': ws_hist.cell(2, 21).value,
    }
    
    print("\n🔍 Historical Transaction Sample:")
    for key, val in hist_transaction.items():
        print(f"   {key:<40} = {val}")
    
    # Look at XML docnomination structure
    print("\n" + "="*100)
    print("ANALYZING DOCNOMINATION FIELD")
    print("="*100)
    
    print("\nChecking first 5 XML records for docnomination patterns:")
    for row in range(2, min(7, ws_xml.max_row + 1)):
        docnomination = ws_xml.cell(row, 7).value  # Column 7 is docnomination
        docprodgroup = ws_xml.cell(row, 16).value   # Column 16 is docprodgroup
        docinformation = ws_xml.cell(row, 8).value  # Column 8 is docinformation
        
        print(f"\n   Row {row}:")
        print(f"      docprodgroup: {docprodgroup}")
        print(f"      docnomination: {docnomination[:80] if docnomination else 'None'}...")
        print(f"      docinformation: {docinformation[:60] if docinformation else 'None'}")
    
    # Check if historical "ოპერაციის შინაარსი" matches docnomination
    print("\n" + "="*100)
    print("FIELD COMPARISON:")
    print("="*100)
    
    comparison = [
        ("თარიღი", "DOCRECDATE / DOCVALUEDATE", "Date field"),
        ("Ref", "DOCKEY", "Document key/reference"),
        ("ოპერაციის იდ", "ENTRIESID", "Entry ID"),
        ("დებეტი", "ENTRYDBAMT", "Debit amount"),
        ("კრედიტი", "ENTRYCRAMT", "Credit amount"),
        ("ოპერაციის შინაარსი", "DOCNOMINATION?", "Operation description"),
        ("დანიშნულება", "ENTRYCOMMENT?", "Purpose/comment"),
        ("დამატებითი ინფორმაცია", "DOCINFORMATION", "Additional info/payment ID"),
        ("ოპერაციის ტიპი", "DOCPRODGROUP", "Operation type (PMD, TRN, COM, etc)"),
    ]
    
    print(f"\n{'Historical Field':<40} | {'XML Field':<30} | {'Notes'}")
    print("-" * 100)
    for hist, xml, note in comparison:
        print(f"{hist:<40} | {xml:<30} | {note}")
    
    wb_hist.close()
    wb_xml.close()

if __name__ == "__main__":
    compare_formats(
        "templates/GE78BG0000000893486000GEL.xlsx",
        "templates/Statement_208678722_reference.xlsx"
    )
