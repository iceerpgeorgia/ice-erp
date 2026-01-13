#!/usr/bin/env python3
"""
Export records that need counteragent account numbers to Excel
User will fill in the accounts and we'll re-import
"""

import openpyxl
import psycopg2

# Database connection
def get_database_url():
    try:
        with open('.env.local', 'r', encoding='utf-8') as f:
            for line in f:
                if line.startswith('DATABASE_URL='):
                    url = line.split('=', 1)[1].strip().strip('"').strip("'")
                    if '?schema=' in url:
                        url = url.split('?schema=')[0]
                    return url
    except Exception as e:
        print(f"❌ Error reading .env.local: {e}")
    return None

def main():
    print("\n" + "="*80)
    print("EXPORT MISSING COUNTERAGENT ACCOUNTS")
    print("="*80 + "\n")

    # Connect to database
    DATABASE_URL = get_database_url()
    if not DATABASE_URL:
        print("❌ Could not get DATABASE_URL")
        return

    try:
        conn = psycopg2.connect(DATABASE_URL)
        cursor = conn.cursor()
        print("✅ Connected to database\n")
    except Exception as e:
        print(f"❌ Database connection failed: {e}")
        return

    # Query records that need accounts
    print("📊 Querying records with missing or corrupted accounts...")
    cursor.execute("""
        SELECT uuid, transaction_date, description, account_currency_amount
        FROM consolidated_bank_accounts
        WHERE counteragent_account_number IS NULL
        ORDER BY transaction_date, id
    """)
    
    records = cursor.fetchall()
    print(f"✅ Found {len(records)} records\n")

    if len(records) == 0:
        print("ℹ️  No records to export!")
        cursor.close()
        conn.close()
        return

    # Create Excel file
    print("📝 Creating Excel file...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Missing Accounts"

    # Headers
    ws['A1'] = 'UUID'
    ws['B1'] = 'Transaction Date'
    ws['C1'] = 'Description'
    ws['D1'] = 'Amount'
    ws['E1'] = 'CounterAgent_Account'
    ws['F1'] = 'Notes'

    # Style headers
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

    # Add data
    for idx, (uuid, trans_date, description, amount) in enumerate(records, start=2):
        ws.cell(row=idx, column=1, value=str(uuid))
        ws.cell(row=idx, column=2, value=trans_date)
        ws.cell(row=idx, column=3, value=description)
        ws.cell(row=idx, column=4, value=float(amount) if amount else None)
        ws.cell(row=idx, column=5, value='')  # Empty for user to fill
        ws.cell(row=idx, column=6, value='')  # Notes

    # Auto-size columns
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 30

    # Save file
    output_file = 'templates/missing_counteragent_accounts.xlsx'
    wb.save(output_file)
    print(f"✅ Saved to: {output_file}\n")

    cursor.close()
    conn.close()

    print("="*80)
    print("SUMMARY")
    print("="*80)
    print(f"  📋 Records exported: {len(records)}")
    print(f"  📄 File: {output_file}")
    print(f"  ℹ️  Note: Raw table is empty, using consolidated table data")
    print(f"  ✏️  Fill in column E (CounterAgent_Account) and save")
    print(f"  🔄 Then run: python scripts/import-filled-accounts.py")
    print("="*80 + "\n")

if __name__ == '__main__':
    main()
