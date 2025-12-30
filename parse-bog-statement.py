import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime

# Parse the XML file
tree = ET.parse('Statement_206598021.xml')
root = tree.getroot()

# Define namespace
ns = {'gemini': 'http://www.mygemini.com/schemas/mygemini'}

# Get header information
header = root.find('HEADER')
account_info = {
    'AcctNo': header.find('AcctNo').text if header.find('AcctNo') is not None else '',
    'ClientName': header.find('ClientName').text if header.find('ClientName') is not None else '',
    'ClientKey': header.find('ClientKey').text if header.find('ClientKey') is not None else '',
    'Ccy': header.find('Ccy').text if header.find('Ccy') is not None else '',
    'Period': header.find('Period').text if header.find('Period') is not None else '',
}

# Create Excel workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "BOG Statement"

# Write header info
ws['A1'] = 'Account Information'
ws['A1'].font = Font(bold=True, size=14)
ws['A2'] = f"Account: {account_info['AcctNo']}"
ws['A3'] = f"Client: {account_info['ClientName']}"
ws['A4'] = f"Client ID: {account_info['ClientKey']}"
ws['A5'] = f"Currency: {account_info['Ccy']}"
ws['A6'] = f"Period: {account_info['Period']}"

# Define columns for transaction details
columns = [
    'DocKey', 'EntriesId', 'DocNo', 'DocRecDate', 'DocValueDate', 
    'EntryDbAmt', 'EntryCrAmt', 'OutBalance',
    'DocNomination', 'DocInformation',
    'DocSenderName', 'DocSenderInn', 'DocSenderAcctNo',
    'DocBenefName', 'DocBenefInn', 'DocBenefAcctNo',
    'DocComment', 'EntryComment',
    'DocProdGroup', 'DocBranch', 'DocDepartment'
]

# Write column headers starting from row 8
header_row = 8
for col_idx, col_name in enumerate(columns, start=1):
    cell = ws.cell(row=header_row, column=col_idx, value=col_name)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')

# Get all DETAIL nodes
details = root.findall('.//DETAIL')

print(f"Found {len(details)} transactions")

# Write transaction data
row = header_row + 1
for detail in details:
    col = 1
    for col_name in columns:
        element = detail.find(col_name)
        value = element.text if element is not None else ''
        ws.cell(row=row, column=col, value=value)
        col += 1
    row += 1

# Auto-adjust column widths
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)
    ws.column_dimensions[column_letter].width = adjusted_width

# Save the workbook
output_file = 'BOG_Statement_Analysis.xlsx'
wb.save(output_file)

print(f"✅ Excel file created: {output_file}")
print(f"📊 Total transactions: {len(details)}")
