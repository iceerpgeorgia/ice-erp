import openpyxl
from openpyxl.utils import get_column_letter
import sys

# Set UTF-8 output encoding
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('handover template.xlsx', data_only=False)
ws = wb['Handover']

print("="*140)
print("DETAILED HANDOVER SHEET - All cells with detailed info\n")

row_count = 0
for row_num in range(1, min(90, ws.max_row + 1)):
    row_has_content = False
    for col_num in range(3, 13):  # Columns C to L
        cell = ws.cell(row=row_num, column=col_num)
        cell_ref = f"{get_column_letter(col_num)}{row_num}"
        
        if cell.value is not None:
            row_has_content = True
            value = cell.value
            
            # Handle array formulas
            if hasattr(cell, 'data_type') and cell.data_type == 'f':
                if hasattr(value, 'text'):
                    formula_text = value.text
                else:
                    formula_text = str(value)
                print(f"{cell_ref:<6} [FORMULA] {formula_text}")
            else:
                val_str = str(value)[:80]
                print(f"{cell_ref:<6} [VALUE]   {val_str}")
    
    if row_has_content:
        row_count += 1

print(f"\nRows with content: {row_count}")

print("\n" + "="*140)
print("\nMERGED CELLS:")
for merged_range in ws.merged_cells.ranges:
    print(f"  {merged_range}")

print("\n" + "="*140)
print("\nCOLUMN WIDTHS:")
for col_num in range(3, 13):
    col_letter = get_column_letter(col_num)
    width = ws.column_dimensions[col_letter].width
    print(f"  Column {col_letter}: {width}")

print("\n" + "="*140)
print("\nKEY INFORMATION:")
print(f"Max row: {ws.max_row}")
print(f"Max column: {ws.max_column}")
