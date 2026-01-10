import openpyxl
import psycopg2
import uuid as uuid_lib
from datetime import datetime

# Database connection
conn = psycopg2.connect(
    host="aws-1-eu-west-1.pooler.supabase.com",
    port=5432,
    database="postgres",
    user="postgres.fojbzghphznbslqwurrm",
    password="fulebimojviT1985%"
)

cursor = conn.cursor()

def import_historical_excel(excel_file, sheet_name='GE78BG0000000893486000GEL'):
    """
    Import historical BOG GEL data from Excel to bog_gel_raw_893486000 table.
    """
    
    print(f"📄 Opening Excel file: {excel_file}")
    wb = openpyxl.load_workbook(excel_file, read_only=True)
    ws = wb[sheet_name]
    
    # Generate import batch ID
    import_batch_id = datetime.now().strftime('%Y%m%d_%H%M%S') + '_historical'
    print(f"🔖 Import batch ID: {import_batch_id}")
    
    # Get headers from first row
    headers = {}
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value:
            headers[cell.value] = col_idx
    
    print(f"📋 Found {len(headers)} columns in Excel")
    
    # Column mapping (Excel header → value)
    def get_col(row, header_name):
        col_idx = headers.get(header_name)
        if col_idx:
            value = row[col_idx - 1].value
            return str(value).strip() if value is not None else None
        return None
    
    # Convert datetime to DD.MM.YYYY format
    def format_date(date_value):
        if date_value is None:
            return None
        if isinstance(date_value, datetime):
            return date_value.strftime('%d.%m.%Y')
        if isinstance(date_value, str):
            # Try to parse if it's already a string date
            try:
                dt = datetime.strptime(date_value, '%Y-%m-%d %H:%M:%S')
                return dt.strftime('%d.%m.%Y')
            except:
                return date_value
        return str(date_value)
    
    # Check for existing duplicates
    print("\n🔍 Checking for existing records...")
    cursor.execute("SELECT dockey, entriesid FROM bog_gel_raw_893486000 WHERE dockey IS NOT NULL AND entriesid IS NOT NULL")
    existing_records = cursor.fetchall()
    existing_keys = set()
    for row in existing_records:
        if row[0] and row[1]:
            # Normalize to string and strip whitespace
            key = (str(row[0]).strip(), str(row[1]).strip())
            existing_keys.add(key)
    print(f"   Found {len(existing_keys)} existing records in database")
    
    # Process rows
    records_to_insert = []
    skipped_missing_keys = 0
    skipped_duplicates = 0
    seen_in_batch = set()  # Track keys seen in current import batch
    
    print(f"\n📊 Processing {ws.max_row - 1} rows from Excel...")
    
    for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
        # Required fields
        dockey = get_col(row, 'Ref')
        entriesid = get_col(row, 'ოპერაციის იდ')
        
        # Skip records without required keys
        if not dockey or not entriesid:
            skipped_missing_keys += 1
            continue
        
        # Normalize keys for comparison
        dockey = str(dockey).strip()
        entriesid = str(entriesid).strip()
        
        # Check for duplicates (in database or in current batch)
        if (dockey, entriesid) in existing_keys or (dockey, entriesid) in seen_in_batch:
            skipped_duplicates += 1
            continue
        
        # Add to seen set
        seen_in_batch.add((dockey, entriesid))
        
        # Extract date
        date_value = row[headers['თარიღი'] - 1].value
        formatted_date = format_date(date_value)
        
        # Build record
        record = {
            'uuid': str(uuid_lib.uuid4()),
            'dockey': dockey,
            'entriesid': entriesid,
            'cancopydocument': None,
            'canviewdocument': None,
            'canprintdocument': None,
            'isreval': None,
            'docnomination': get_col(row, 'ოპერაციის შინაარსი'),
            'docinformation': get_col(row, 'დამატებითი ინფორმაცია'),
            'docsrcamt': None,
            'docsrcccy': None,
            'docdstamt': None,
            'docdstccy': None,
            'docrecdate': formatted_date,
            'docbranch': None,
            'docdepartment': None,
            'docprodgroup': get_col(row, 'ოპერაციის ტიპი'),
            'docno': get_col(row, 'საბუთის N'),
            'docvaluedate': formatted_date,
            'docsendername': get_col(row, 'გამგზავნის დასახელება'),
            'docsenderinn': get_col(row, 'გამგზავნის საიდენტიფიკაციო კოდი'),
            'docsenderacctno': get_col(row, 'გამგზავნის ანგარიშის ნომერი'),
            'docsenderbic': get_col(row, 'გამგზავნი ბანკის კოდი'),
            'docactualdate': None,
            'doccoracct': get_col(row, 'მოკორესპოდენტო ანგარიში'),
            'doccorbic': None,
            'doccorbankname': None,
            'doccomment': None,
            'ccyrate': None,
            'entrypdate': None,
            'entrydocno': None,
            'entrylacct': None,
            'entrylacctold': None,
            'entrydbamt': get_col(row, 'დებეტი'),
            'entrydbamtbase': None,
            'entrycramt': get_col(row, 'კრედიტი'),
            'entrycramtbase': None,
            'outbalance': get_col(row, 'ნაშთი ოპერაციის ბოლოს'),
            'entryamtbase': None,
            'entrycomment': get_col(row, 'დანიშნულება'),
            'entrydepartment': None,
            'entryacctpoint': None,
            'docsenderbicname': get_col(row, 'გამგზავნი ბანკის დასახელება'),
            'docbenefname': get_col(row, 'მიმღების დასახელება'),
            'docbenefinn': get_col(row, 'მიმღების საიდენტიფიკაციო კოდი'),
            'docbenefacctno': get_col(row, 'მიმღების ანგარიშის ნომერი'),
            'docbenefbic': get_col(row, 'მიმღები ბანკის კოდი'),
            'docbenefbicname': get_col(row, 'მიმღები ბანკის დასახელება'),
            'docpayername': None,
            'docpayerinn': None,
            'import_batch_id': import_batch_id,
            'is_processed': False
        }
        
        records_to_insert.append(record)
        
        # Progress update every 1000 records
        if len(records_to_insert) % 1000 == 0:
            print(f"   Processed {len(records_to_insert)} records...")
    
    wb.close()
    
    print(f"\n📊 Processing Summary:")
    print(f"   ✅ Valid records to insert: {len(records_to_insert)}")
    print(f"   ⚠️  Skipped (missing keys): {skipped_missing_keys}")
    print(f"   🔄 Skipped (duplicates): {skipped_duplicates}")
    
    if len(records_to_insert) == 0:
        print("\n⚠️  No new records to insert!")
        return
    
    # Insert records
    print(f"\n💾 Inserting {len(records_to_insert)} records into bog_gel_raw_893486000...")
    
    insert_query = """
        INSERT INTO bog_gel_raw_893486000 (
            uuid, dockey, entriesid, cancopydocument, canviewdocument, canprintdocument,
            isreval, docnomination, docinformation, docsrcamt, docsrcccy, docdstamt, docdstccy,
            docrecdate, docbranch, docdepartment, docprodgroup, docno, docvaluedate,
            docsendername, docsenderinn, docsenderacctno, docsenderbic, docactualdate,
            doccoracct, doccorbic, doccorbankname, doccomment, ccyrate, entrypdate,
            entrydocno, entrylacct, entrylacctold, entrydbamt, entrydbamtbase,
            entrycramt, entrycramtbase, outbalance, entryamtbase, entrycomment,
            entrydepartment, entryacctpoint, docsenderbicname, docbenefname, docbenefinn,
            docbenefacctno, docbenefbic, docbenefbicname, docpayername, docpayerinn,
            import_batch_id, is_processed
        ) VALUES (
            %(uuid)s, %(dockey)s, %(entriesid)s, %(cancopydocument)s, %(canviewdocument)s, 
            %(canprintdocument)s, %(isreval)s, %(docnomination)s, %(docinformation)s, 
            %(docsrcamt)s, %(docsrcccy)s, %(docdstamt)s, %(docdstccy)s, %(docrecdate)s,
            %(docbranch)s, %(docdepartment)s, %(docprodgroup)s, %(docno)s, %(docvaluedate)s,
            %(docsendername)s, %(docsenderinn)s, %(docsenderacctno)s, %(docsenderbic)s,
            %(docactualdate)s, %(doccoracct)s, %(doccorbic)s, %(doccorbankname)s,
            %(doccomment)s, %(ccyrate)s, %(entrypdate)s, %(entrydocno)s, %(entrylacct)s,
            %(entrylacctold)s, %(entrydbamt)s, %(entrydbamtbase)s, %(entrycramt)s,
            %(entrycramtbase)s, %(outbalance)s, %(entryamtbase)s, %(entrycomment)s,
            %(entrydepartment)s, %(entryacctpoint)s, %(docsenderbicname)s, %(docbenefname)s,
            %(docbenefinn)s, %(docbenefacctno)s, %(docbenefbic)s, %(docbenefbicname)s,
            %(docpayername)s, %(docpayerinn)s, %(import_batch_id)s, %(is_processed)s
        )
    """
    
    inserted_count = 0
    batch_size = 500
    
    for i in range(0, len(records_to_insert), batch_size):
        batch = records_to_insert[i:i+batch_size]
        for record in batch:
            cursor.execute(insert_query, record)
            inserted_count += 1
        
        conn.commit()
        print(f"   Inserted {inserted_count}/{len(records_to_insert)} records...")
    
    print(f"\n✅ Successfully inserted {inserted_count} records!")
    print(f"📋 Batch ID: {import_batch_id}")
    print(f"⚠️  Note: Records marked as is_processed=false (not processed to consolidated table yet)")

if __name__ == "__main__":
    try:
        import_historical_excel("templates/GE78BG0000000893486000GEL.xlsx")
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
    finally:
        cursor.close()
        conn.close()
