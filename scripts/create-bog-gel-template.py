import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_bog_gel_raw_template():
    """
    Create Excel template for bog_gel_raw_893486000 table import.
    Only includes user-fillable columns (excludes id, uuid, timestamps, and flags).
    """
    
    # Define columns from XML import (matching import-bog-statement.py)
    columns = [
        # Document Key Fields
        'dockey',              # Document Key (REQUIRED)
        'entriesid',           # Entry ID (REQUIRED)
        
        # Document Metadata
        'cancopydocument',
        'canviewdocument', 
        'canprintdocument',
        'isreval',
        
        # Document Information
        'docnomination',        # Description
        'docinformation',       # Payment ID or reference
        'docsrcamt',           # Source amount
        'docsrcccy',           # Source currency
        'docdstamt',           # Destination amount  
        'docdstccy',           # Destination currency
        'docrecdate',          # Recording date (format: DD.MM.YYYY)
        'docbranch',           # Branch
        'docdepartment',       # Department
        'docprodgroup',        # Product Group (COM, CCO, TRN, etc.)
        'docno',               # Document number
        'docvaluedate',        # Value date (format: DD.MM.YYYY) (REQUIRED)
        
        # Sender Information
        'docsendername',       # Sender name
        'docsenderinn',        # Sender INN
        'docsenderacctno',     # Sender account number
        'docsenderbic',        # Sender BIC
        'docsenderbicname',    # Sender BIC name
        
        # Beneficiary Information
        'docbenefname',        # Beneficiary name
        'docbenefinn',         # Beneficiary INN
        'docbenefacctno',      # Beneficiary account number
        'docbenefbic',         # Beneficiary BIC
        'docbenefbicname',     # Beneficiary BIC name
        
        # Payer Information
        'docpayername',        # Payer name
        'docpayerinn',         # Payer INN
        
        # Additional Document Fields
        'docactualdate',       # Actual date
        'doccoracct',          # Correspondent account
        'doccorbic',           # Correspondent BIC
        'doccorbankname',      # Correspondent bank name
        'doccomment',          # Document comment
        
        # Entry Information
        'ccyrate',             # Currency rate
        'entrypdate',          # Entry posting date
        'entrydocno',          # Entry document number
        'entrylacct',          # Entry local account
        'entrylacctold',       # Entry local account old
        'entrydbamt',          # Entry debit amount
        'entrydbamtbase',      # Entry debit amount base
        'entrycramt',          # Entry credit amount
        'entrycramtbase',      # Entry credit amount base
        'outbalance',          # Outstanding balance
        'entryamtbase',        # Entry amount base
        'entrycomment',        # Entry comment
        'entrydepartment',     # Entry department
        'entryacctpoint',      # Entry account point
    ]
    
    # Column descriptions/help text
    descriptions = {
        'dockey': 'REQUIRED - Document Key from bank statement',
        'entriesid': 'REQUIRED - Entry ID from bank statement', 
        'docvaluedate': 'REQUIRED - Transaction date (format: DD.MM.YYYY or DD.MM.YY)',
        'docrecdate': 'Recording date (format: DD.MM.YYYY or DD.MM.YY)',
        'docnomination': 'Transaction description/purpose',
        'docinformation': 'Payment ID or reference number',
        'docprodgroup': 'Product Group: COM, CCO, TRN, PMC, PMD, etc.',
        'docsenderinn': 'Sender Tax ID / INN',
        'docbenefinn': 'Beneficiary Tax ID / INN',
        'entrycramt': 'Credit amount (incoming)',
        'entrydbamt': 'Debit amount (outgoing)',
        'docsrcccy': 'Source currency code (USD, EUR, GEL, etc.)',
        'docdstccy': 'Destination currency code',
    }
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOG_GEL_RAW_Import"
    
    # Set column headers with styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.value = col_name
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Set column width
        ws.column_dimensions[get_column_letter(idx)].width = 18
    
    # Add description row with help text
    for idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=2, column=idx)
        cell.value = descriptions.get(col_name, "")
        cell.font = Font(italic=True, size=9, color="666666")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    ws.row_dimensions[2].height = 30
    
    # Add sample data row
    sample_data = {
        'dockey': '31117184347',
        'entriesid': '108518127520',
        'docvaluedate': '25.12.2025',
        'docrecdate': '25.12.2025',
        'docnomination': 'მექნიკური მონტაჟის ღირებულება',
        'docinformation': 'a6042a_48_b0e8b1',
        'docprodgroup': 'COM',
        'entrycramt': '15000.00',
        'entrydbamt': '0.00',
        'docsrcccy': 'GEL',
        'docdstccy': 'GEL',
        'docsenderinn': '45001031265',
        'docsendername': 'SENDER COMPANY LLC',
        'docbenefinn': '',
        'docbenefname': '',
    }
    
    for idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=idx)
        cell.value = sample_data.get(col_name, '')
        cell.font = Font(color="999999", italic=True)
    
    # Freeze top 2 rows
    ws.freeze_panes = "A3"
    
    # Save template
    import time
    timestamp = int(time.time())
    filename = f'templates/bog_gel_raw_893486000_import_template_{timestamp}.xlsx'
    wb.save(filename)
    print(f"✅ Template created: {filename}")
    print(f"📋 Columns: {len(columns)}")
    print(f"⚠️  REQUIRED fields: dockey, entriesid, docvaluedate")
    print(f"💡 Note: id, uuid, timestamps, is_processed, and import_batch_id will be auto-generated on import")

if __name__ == "__main__":
    create_bog_gel_raw_template()
