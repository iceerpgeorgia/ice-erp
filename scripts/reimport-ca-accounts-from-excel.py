import openpyxl
import psycopg2
from datetime import datetime
import re

# Database connection - read from .env.local
def get_database_url():
    try:
        with open('.env.local', 'r', encoding='utf-8') as f:
            for line in f:
                if line.startswith('DATABASE_URL='):
                    url = line.split('=', 1)[1].strip().strip('"').strip("'")
                    # Remove schema parameter if present (psycopg2 doesn't support it)
                    if '?schema=' in url:
                        url = url.split('?schema=')[0]
                    return url
    except Exception as e:
        print(f"❌ Error reading .env.local: {e}")
    return None

DATABASE_URL = get_database_url()
if not DATABASE_URL:
    print("❌ DATABASE_URL not found in .env.local")
    exit(1)

conn = psycopg2.connect(DATABASE_URL)
cursor = conn.cursor()

def reimport_ca_accounts(excel_file, sheet_name='GE78BG0000000893486000GEL'):
    """
    Re-import counteragent account numbers from Excel to update consolidated_bank_accounts.
    Matches by DocKey (Ref) and EntriesId (ოპერაციის იდ).
    """
    
    print(f"📄 Opening Excel file: {excel_file}\n")
    wb = openpyxl.load_workbook(excel_file, read_only=True)
    ws = wb[sheet_name]
    
    # Get headers from first row
    headers = {}
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value:
            headers[cell.value] = col_idx
    
    print(f"📋 Found {len(headers)} columns in Excel")
    print(f"   Required columns: Ref, ოპერაციის იდ, გამგზავნის ანგარიშის ნომერი, ბენეფიციარის ანგარიშის ნომერი, დებეტი\n")
    
    # Column mapping
    def get_col(row, header_name):
        col_idx = headers.get(header_name)
        if col_idx:
            value = row[col_idx - 1].value
            return str(value).strip() if value is not None else None
        return None
    
    # Convert scientific notation to proper integer string
    def convert_from_scientific(value):
        if not value:
            return None
        
        value_str = str(value).strip()
        
        # Check if it's in scientific notation (contains 'e' or 'E')
        if 'e' in value_str.lower():
            try:
                # Convert to float then to int to remove scientific notation
                numeric_value = float(value_str)
                # Convert to integer (removing decimal point)
                int_value = int(numeric_value)
                return str(int_value)
            except (ValueError, OverflowError):
                # If conversion fails, return original
                print(f"   ⚠️  Could not convert scientific notation: {value_str}")
                return value_str
        
        return value_str
    
    print(f"🔄 Processing {ws.max_row - 1} rows from Excel...\n")
    
    updated_count = 0
    not_found_count = 0
    no_account_count = 0
    error_count = 0
    
    for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
        try:
            # Get key fields
            dockey = get_col(row, 'Ref')
            entriesid = get_col(row, 'ოპერაციის იდ')
            
            if not dockey or not entriesid:
                continue
            
            # Normalize keys
            dockey = str(dockey).strip()
            entriesid = str(entriesid).strip()
            
            # Get account numbers and convert from scientific notation if needed
            sender_account = get_col(row, 'გამგზავნის ანგარიშის ნომერი')
            benef_account = get_col(row, 'ბენეფიციარის ანგარიშის ნომერი')
            debit = get_col(row, 'დებეტი')
            
            # Convert from scientific notation
            sender_account = convert_from_scientific(sender_account)
            benef_account = convert_from_scientific(benef_account)
            
            # Determine which account is the counteragent's
            # If debit is NULL/empty -> incoming payment -> sender is counteragent
            # If debit has value -> outgoing payment -> beneficiary is counteragent
            counteragent_account = None
            
            if not debit or debit == 'None' or debit == '':
                # Incoming - sender is counteragent
                counteragent_account = sender_account
            else:
                # Outgoing - beneficiary is counteragent
                counteragent_account = benef_account
            
            if not counteragent_account:
                no_account_count += 1
                continue
            
            # Find matching record in consolidated_bank_accounts via raw table
            cursor.execute("""
                SELECT c.id, c.counteragent_account_number
                FROM consolidated_bank_accounts c
                JOIN bog_gel_raw_893486000 r ON c.raw_record_uuid = r.uuid
                WHERE r.dockey = %s AND r.entriesid = %s
            """, (dockey, entriesid))
            
            result = cursor.fetchone()
            
            if not result:
                not_found_count += 1
                continue
            
            consolidated_id = result[0]
            current_account = result[1]
            
            # Only update if different or NULL
            if current_account != counteragent_account:
                cursor.execute("""
                    UPDATE consolidated_bank_accounts
                    SET counteragent_account_number = %s, updated_at = NOW()
                    WHERE id = %s
                """, (counteragent_account, consolidated_id))
                
                updated_count += 1
                
                if updated_count % 100 == 0:
                    conn.commit()  # Commit every 100 records
                    print(f"   ✅ Updated {updated_count}...")
            
        except Exception as e:
            error_count += 1
            if error_count <= 5:
                print(f"   ❌ Error at row {row_num}: {str(e)}")
    
    # Final commit
    conn.commit()
    
    print(f"\n{'='*80}")
    print(f"SUMMARY")
    print(f"{'='*80}")
    print(f"  ✅ Updated: {updated_count}")
    print(f"  ⚠️  Not found in DB: {not_found_count}")
    print(f"  ℹ️  No account in Excel: {no_account_count}")
    print(f"  ❌ Errors: {error_count}")
    print(f"{'='*80}\n")
    
    wb.close()

if __name__ == "__main__":
    try:
        reimport_ca_accounts("templates/GE78BG0000000893486000GEL.xlsx")
        print("✅ Re-import completed successfully!")
    except Exception as e:
        print(f"❌ Failed: {str(e)}")
        conn.rollback()
    finally:
        cursor.close()
        conn.close()
